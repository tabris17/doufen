# encoding: utf-8
import ast
import logging

from openpyxl import Workbook

import db
from .handlers import BaseRequestHandler


class Index(BaseRequestHandler):
    """
    导出主页
    """
    def get(self):
        self.render('exports.html')

    def post(self):
        filename = self.get_argument('filename')
        items = set(self.get_argument('items').split(','))
        workbook = Workbook()
        workbook.remove(workbook.active)
        for item in items:
            if item == 'friend':
                self._export_friend(workbook)
            elif item == 'movie':
                self._export_movie(workbook)
            elif item == 'music':
                self._export_music(workbook)
            elif item == 'book':
                self._export_book(workbook)
            elif item == 'broadcast':
                self._export_broadcast(workbook)
                
        workbook.save(filename)
        self.write('OK')

    def _export_movie(self, workbook):
        def fill_thead(ws):
            ws.cell(row=1, column=1, value='ID')
            ws.cell(row=1, column=2, value='标题')
            ws.cell(row=1, column=3, value='又名')
            ws.cell(row=1, column=4, value='导演')
            ws.cell(row=1, column=5, value='编剧')
            ws.cell(row=1, column=6, value='主演')
            ws.cell(row=1, column=7, value='类型')
            ws.cell(row=1, column=8, value='国家地区')
            ws.cell(row=1, column=9, value='语言')
            ws.cell(row=1, column=10, value='首播')
            ws.cell(row=1, column=11, value='集数')
            ws.cell(row=1, column=12, value='单集片长')
            ws.cell(row=1, column=13, value='豆瓣评分')
            ws.cell(row=1, column=14, value='评分人数')
            ws.cell(row=1, column=15, value='链接')
            ws.cell(row=1, column=16, value='我的评价')
            ws.cell(row=1, column=17, value='评价时间')
            ws.cell(row=1, column=18, value='标签')

        def fill_my_movie(ws, row, my_movie):
            movie = my_movie.movie
            attrs = ast.literal_eval(movie.attrs)
            ws.cell(row=row, column=1, value=movie.douban_id)
            ws.cell(row=row, column=2, value=movie.title)
            ws.cell(row=row, column=3, value=movie.alt_title)
            if 'director' in attrs and attrs['director']:
                ws.cell(row=row, column=4, value=' / '.join(attrs['director']))
            if 'writer' in attrs and attrs['writer']:
                ws.cell(row=row, column=5, value=' / '.join(attrs['writer']))
            if 'cast' in attrs and attrs['cast']:
                ws.cell(row=row, column=6, value=' / '.join(attrs['cast']))
            if 'movie_type' in attrs and attrs['movie_type']:
                ws.cell(row=row, column=7, value=' / '.join(attrs['movie_type']))
            if 'country' in attrs and attrs['country']:
                ws.cell(row=row, column=8, value=' / '.join(attrs['country']))
            if 'language' in attrs and attrs['language']:
                ws.cell(row=row, column=9, value=' / '.join(attrs['language']))
            if 'pubdate' in attrs and attrs['pubdate']:
                ws.cell(row=row, column=10, value=' / '.join(attrs['pubdate']))
            if 'episodes' in attrs and attrs['episodes']:
                ws.cell(row=row, column=11, value=' / '.join(attrs['episodes']))
            if 'movie_duration' in attrs and attrs['movie_duration']:
                ws.cell(row=row, column=12, value=' / '.join(attrs['movie_duration']))
            if movie.rating:
                rating = ast.literal_eval(movie.rating)
                ws.cell(row=row, column=13, value=rating['average'])
                ws.cell(row=row, column=14, value=rating['numRaters'])
            ws.cell(row=row, column=15, value=movie.alt)
            if my_movie.rating:
                my_rating = ast.literal_eval(my_movie.rating)
                ws.cell(row=row, column=16, value=my_rating['value'])
            ws.cell(row=row, column=17, value=my_movie.create_time)
            if my_movie.tags:
                tags = ast.literal_eval(my_movie.tags)
                if len(tags):
                    ws.cell(row=row, column=18, value=' / '.join(tags))

        worksheet = workbook.create_sheet('看过的电影')
        fill_thead(worksheet)
        query = db.MyMovie.select(db.MyMovie, db.Movie).join(
            db.Movie, on=db.MyMovie.movie
        ).where(db.MyMovie.user == self.get_current_user(), db.MyMovie.status == 'done').order_by(db.MyMovie.id.desc())
        row_num = 2
        for row in query:
            fill_my_movie(worksheet, row_num, row)
            row_num += 1

        worksheet = workbook.create_sheet('想看的电影')
        fill_thead(worksheet)
        query = db.MyMovie.select(db.MyMovie, db.Movie).join(
            db.Movie, on=db.MyMovie.movie
        ).where(db.MyMovie.user == self.get_current_user(), db.MyMovie.status == 'wish').order_by(db.MyMovie.id.desc())
        row_num = 2
        for row in query:
            fill_my_movie(worksheet, row_num, row)
            row_num += 1

        worksheet = workbook.create_sheet('在看的电视剧')
        fill_thead(worksheet)
        query = db.MyMovie.select(db.MyMovie, db.Movie).join(
            db.Movie, on=db.MyMovie.movie
        ).where(db.MyMovie.user == self.get_current_user(), db.MyMovie.status == 'doing').order_by(db.MyMovie.id.desc())
        row_num = 2
        for row in query:
            fill_my_movie(worksheet, row_num, row)
            row_num += 1


    def _export_music(self, workbook):
        def fill_thead(ws):
            ws.cell(row=1, column=1, value='ID')
            ws.cell(row=1, column=2, value='标题')
            ws.cell(row=1, column=3, value='又名')
            ws.cell(row=1, column=4, value='表演者')
            ws.cell(row=1, column=5, value='专辑类型')
            ws.cell(row=1, column=6, value='介质')
            ws.cell(row=1, column=7, value='发行时间')
            ws.cell(row=1, column=8, value='出版者')
            ws.cell(row=1, column=9, value='唱片数')
            ws.cell(row=1, column=10, value='豆瓣评分')
            ws.cell(row=1, column=11, value='评分人数')
            ws.cell(row=1, column=12, value='链接')
            ws.cell(row=1, column=13, value='我的评价')
            ws.cell(row=1, column=14, value='评价时间')
            ws.cell(row=1, column=15, value='标签')

        def fill_my_music(ws, row, my_music):
            music = my_music.music
            attrs = ast.literal_eval(music.attrs)
            ws.cell(row=row, column=1, value=music.douban_id)
            ws.cell(row=row, column=2, value=music.title)
            ws.cell(row=row, column=3, value=music.alt_title)
            if 'singer' in attrs and attrs['singer']:
                ws.cell(row=row, column=4, value=' / '.join(attrs['singer']))
            if 'version' in attrs and attrs['version']:
                ws.cell(row=row, column=5, value=' / '.join(attrs['version']))
            if 'media' in attrs and attrs['media']:
                ws.cell(row=row, column=6, value=' / '.join(attrs['media']))
            if 'pubdate' in attrs and attrs['pubdate']:
                ws.cell(row=row, column=7, value=' / '.join(attrs['pubdate']))
            if 'publisher' in attrs and attrs['publisher']:
                ws.cell(row=row, column=8, value=' / '.join(attrs['publisher']))
            if 'discs' in attrs and attrs['discs']:
                ws.cell(row=row, column=9, value=' / '.join(attrs['discs']))
            if music.rating:
                rating = ast.literal_eval(music.rating)
                ws.cell(row=row, column=10, value=rating['average'])
                ws.cell(row=row, column=11, value=rating['numRaters'])
            ws.cell(row=row, column=12, value=music.alt)
            if my_music.rating:
                my_rating = ast.literal_eval(my_music.rating)
                ws.cell(row=row, column=13, value=my_rating['value'])
            ws.cell(row=row, column=14, value=my_music.create_time)
            if my_music.tags:
                tags = ast.literal_eval(my_music.tags)
                if len(tags):
                    ws.cell(row=row, column=15, value=' / '.join(tags))
                

        worksheet = workbook.create_sheet('听过的唱片')
        fill_thead(worksheet)
        query = db.MyMusic.select(db.MyMusic, db.Music).join(
            db.Music, on=db.MyMusic.music
        ).where(db.MyMusic.user == self.get_current_user(), db.MyMusic.status == 'done').order_by(db.MyMusic.id.desc())
        row_num = 2
        for row in query:
            fill_my_music(worksheet, row_num, row)
            row_num += 1

        worksheet = workbook.create_sheet('想听的唱片')
        fill_thead(worksheet)
        query = db.MyMusic.select(db.MyMusic, db.Music).join(
            db.Music, on=db.MyMusic.music
        ).where(db.MyMusic.user == self.get_current_user(), db.MyMusic.status == 'wish').order_by(db.MyMusic.id.desc())
        row_num = 2
        for row in query:
            fill_my_music(worksheet, row_num, row)
            row_num += 1


    def _export_book(self, workbook):
        def fill_thead(ws):
            ws.cell(row=1, column=1, value='ID')
            ws.cell(row=1, column=2, value='标题')
            ws.cell(row=1, column=3, value='副标题')
            ws.cell(row=1, column=4, value='又名')
            ws.cell(row=1, column=5, value='作者')
            ws.cell(row=1, column=6, value='译者')
            ws.cell(row=1, column=7, value='出版社')
            ws.cell(row=1, column=8, value='原作名')
            ws.cell(row=1, column=9, value='出版日期')
            ws.cell(row=1, column=10, value='ISBN')
            ws.cell(row=1, column=11, value='价格')
            ws.cell(row=1, column=12, value='页数')
            ws.cell(row=1, column=13, value='装帧')
            ws.cell(row=1, column=14, value='豆瓣评分')
            ws.cell(row=1, column=15, value='评分人数')
            ws.cell(row=1, column=16, value='链接')
            ws.cell(row=1, column=17, value='我的评价')
            ws.cell(row=1, column=18, value='评价时间')
            ws.cell(row=1, column=19, value='标签')

        def fill_my_book(ws, row, my_book):
            book = my_book.book
            ws.cell(row=row, column=1, value=book.douban_id)
            ws.cell(row=row, column=2, value=book.title)
            ws.cell(row=row, column=3, value=book.subtitle)
            ws.cell(row=row, column=4, value=book.alt_title)
            if book.author:
                author = ast.literal_eval(book.author)
                if len(author):
                    ws.cell(row=row, column=5, value=' / '.join(author))
            if book.translator:
                translator = ast.literal_eval(book.translator)
                if len(translator):
                    ws.cell(row=row, column=6, value=' / '.join(translator))
            ws.cell(row=row, column=7, value=book.publisher)
            ws.cell(row=row, column=8, value=book.origin_title)
            ws.cell(row=row, column=9, value=book.pubdate)
            ws.cell(row=row, column=10, value='{0} / {1}'.format(book.isbn10, book.isbn13))
            ws.cell(row=row, column=11, value=book.price)
            ws.cell(row=row, column=12, value=book.pages)
            ws.cell(row=row, column=13, value=book.binding)
            if book.rating:
                rating = ast.literal_eval(book.rating)
                ws.cell(row=row, column=14, value=rating['average'])
                ws.cell(row=row, column=15, value=rating['numRaters'])
            ws.cell(row=row, column=16, value=book.alt)
            if my_book.rating:
                my_rating = ast.literal_eval(my_book.rating)
                ws.cell(row=row, column=17, value=my_rating['value'])
            ws.cell(row=row, column=18, value=my_book.create_time)
            if my_book.tags:
                tags = ast.literal_eval(my_book.tags)
                if len(tags):
                    ws.cell(row=row, column=19, value=' / '.join(tags))

        worksheet = workbook.create_sheet('读过的书')
        fill_thead(worksheet)
        query = db.MyBook.select(db.MyBook, db.Book).join(
            db.Book, on=db.MyBook.book
        ).where(db.MyBook.user == self.get_current_user(), db.MyBook.status == 'done').order_by(db.MyBook.id.desc())
        row_num = 2
        for row in query:
            fill_my_book(worksheet, row_num, row)
            row_num += 1

        worksheet = workbook.create_sheet('想读的书')
        fill_thead(worksheet)
        query = db.MyBook.select(db.MyBook, db.Book).join(
            db.Book, on=db.MyBook.book
        ).where(db.MyBook.user == self.get_current_user(), db.MyBook.status == 'wish').order_by(db.MyBook.id.desc())
        row_num = 2
        for row in query:
            fill_my_book(worksheet, row_num, row)
            row_num += 1

        worksheet = workbook.create_sheet('在读的书')
        fill_thead(worksheet)
        query = db.MyBook.select(db.MyBook, db.Book).join(
            db.Book, on=db.MyBook.book
        ).where(db.MyBook.user == self.get_current_user(), db.MyBook.status == 'doing').order_by(db.MyBook.id.desc())
        row_num = 2
        for row in query:
            fill_my_book(worksheet, row_num, row)
            row_num += 1


    def _export_friend(self, workbook):
        def fill_thead(ws):
            ws.cell(row=1, column=1, value='ID')
            ws.cell(row=1, column=2, value='域名')
            ws.cell(row=1, column=3, value='名号')
            ws.cell(row=1, column=4, value='注册时间')
            ws.cell(row=1, column=5, value='常居地')
            ws.cell(row=1, column=6, value='用户主页')

        def fill_user(ws, row, user):
            ws.cell(row=row, column=1, value=user.douban_id)
            ws.cell(row=row, column=2, value=user.unique_name)
            ws.cell(row=row, column=3, value=user.name)
            ws.cell(row=row, column=4, value=user.created)
            ws.cell(row=row, column=5, value=user.loc_name)
            ws.cell(row=row, column=6, value=user.alt)

        worksheet = workbook.create_sheet('我关注的人')
        fill_thead(worksheet)
        query = db.Following.select(db.Following, db.User).join(
            db.User, on=db.Following.following_user
        ).where(db.Following.user == self.get_current_user()).order_by(db.Following.id.desc())
        row_num = 2
        for row in query:
            fill_user(worksheet, row_num, row.following_user)
            row_num += 1
        
        worksheet = workbook.create_sheet('关注我的人')
        fill_thead(worksheet)
        query = db.Follower.select(db.Follower, db.User).join(
            db.User, on=db.Follower.follower
        ).where(db.Follower.user == self.get_current_user()).order_by(db.Follower.id.desc())
        row_num = 2
        for row in query:
            fill_user(worksheet, row_num, row.follower)
            row_num += 1

        worksheet = workbook.create_sheet('黑名单')
        fill_thead(worksheet)
        query = db.BlockUser.select(db.BlockUser, db.User).join(
            db.User, on=db.BlockUser.block_user
        ).where(db.BlockUser.user == self.get_current_user()).order_by(db.BlockUser.id.desc())
        row_num = 2
        for row in query:
            fill_user(worksheet, row_num, row.block_user)
            row_num += 1


    def _export_broadcast(self, workbook):
        def fill_thead(ws):
            ws.cell(row=1, column=1, value='ID')
            ws.cell(row=1, column=2, value='用户ID')
            ws.cell(row=1, column=3, value='用户域名')
            ws.cell(row=1, column=4, value='用户名号')
            ws.cell(row=1, column=5, value='文字内容')
            ws.cell(row=1, column=6, value='完整内容')
            ws.cell(row=1, column=7, value='地址')
            ws.cell(row=1, column=8, value='发表时间')
            ws.cell(row=1, column=9, value='回应')
            ws.cell(row=1, column=10, value='推荐')
            ws.cell(row=1, column=11, value='转播')

        worksheet = workbook.create_sheet('广播')
        fill_thead(worksheet)
        query = db.Timeline.select(
            db.Timeline, 
            db.Broadcast, 
            db.User
        ).join(db.Broadcast).join(db.User, db.JOIN.LEFT_OUTER, on=db.Timeline.broadcast.user).where(db.Timeline.user == self.get_current_user()).order_by(db.Timeline.id.desc())
        row_num = 2
        for row in query:
            broadcast = row.broadcast
            worksheet.cell(row=row_num, column=1, value=broadcast.douban_id)
            if broadcast.user:
                worksheet.cell(row=row_num, column=2, value=broadcast.user.douban_id)
                worksheet.cell(row=row_num, column=3, value=broadcast.user.unique_name)
                worksheet.cell(row=row_num, column=4, value=broadcast.user.name)
            worksheet.cell(row=row_num, column=5, value=broadcast.blockquote)
            worksheet.cell(row=row_num, column=6, value=broadcast.content)
            worksheet.cell(row=row_num, column=7, value=broadcast.status_url)
            worksheet.cell(row=row_num, column=8, value=broadcast.created)
            worksheet.cell(row=row_num, column=9, value=broadcast.comments_count)
            worksheet.cell(row=row_num, column=10, value=broadcast.like_count)
            worksheet.cell(row=row_num, column=11, value=broadcast.reshared_count)
            row_num += 1